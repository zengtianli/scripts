#!/usr/bin/env python3
"""
为 Excel 文件中重复的企业编码生成唯一用户编码
"""

import sys
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook


def generate_user_codes(file_path: str, output_path: str = None):
    """
    为重复的企业编码生成唯一用户编码

    Args:
        file_path: 输入 Excel 文件路径
        output_path: 输出文件路径，默认为原文件名_已编号.xlsx
    """
    # 加载工作簿
    wb = load_workbook(file_path)
    ws = wb.active

    # 查找"企业编码"列
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    try:
        code_col_idx = header_row.index("企业编码") + 1  # openpyxl 列索引从1开始
    except ValueError:
        print("错误：未找到'企业编码'列")
        return

    # 统计企业编码出现次数
    codes = []
    for row in ws.iter_rows(min_row=2, min_col=code_col_idx, max_col=code_col_idx, values_only=True):
        if row[0]:  # 跳过空值
            codes.append(row[0])

    code_counts = Counter(codes)

    # 记录每个重复编码的当前序号
    code_sequence = {}

    # 在"企业编码"列后插入"用户编码"列
    ws.insert_cols(code_col_idx + 1)
    ws.cell(row=1, column=code_col_idx + 1, value="用户编码")

    # 生成用户编码
    for row_idx in range(2, ws.max_row + 1):
        enterprise_code = ws.cell(row=row_idx, column=code_col_idx).value

        if not enterprise_code:
            continue

        # 所有编码都添加序号后缀
        if enterprise_code not in code_sequence:
            code_sequence[enterprise_code] = 1
        else:
            code_sequence[enterprise_code] += 1

        user_code = f"{enterprise_code}B{code_sequence[enterprise_code]:04d}"
        ws.cell(row=row_idx, column=code_col_idx + 1, value=user_code)

    # 保存文件
    if output_path is None:
        input_path = Path(file_path)
        output_path = input_path.parent / f"{input_path.stem}_已编号{input_path.suffix}"

    wb.save(str(output_path))
    print(f"✓ 处理完成，已保存到：{output_path}")
    print(f"  - 共处理 {len(codes)} 条记录")
    print(f"  - 重复编码数：{sum(1 for count in code_counts.values() if count > 1)}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python3 xlsx_encode_duplicates.py <输入文件> [输出文件]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    generate_user_codes(input_file, output_file)
