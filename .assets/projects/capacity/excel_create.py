#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ============================================================
# 脚本名称: create_xlsm.py
# 功能描述: 创建带 VBA 宏的 Excel 模板文件 (.xlsm)
# 来源工单: 水利公司需求
# 创建日期: 2025-01-13
# 作者: 开发部
# ============================================================
"""
创建纳污能力计算模板

使用方式:
    python create_xlsm.py           # 创建带示例数据的模板
    python create_xlsm.py --empty   # 创建空模板
"""

import argparse
from pathlib import Path
import pandas as pd
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    exit(1)


def create_template(output_path: Path, with_sample: bool = True):
    """创建 Excel 模板"""
    
    wb = Workbook()
    
    # ========== Sheet 1: 功能区基础信息 ==========
    ws1 = wb.active
    ws1.title = "功能区基础信息"
    
    # 表头
    headers1 = ["功能区", "名称", "水质类别", "Cs", "C0", "河段长度L(m)", 
                "衰减系数K(1/s)", "不均匀系数b", "a", "β"]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 示例数据
    if with_sample:
        sample_zones = [
            ["QT-153", "源头段", "II", 0.5, 0.02, 1000, 0.001, 0.8, 0.3, 0.5],
            ["QT-154", "上游段", "II", 0.5, 0.02, 1500, 0.001, 0.8, 0.3, 0.5],
            ["QT-155", "中游段", "III", 1.0, 0.04, 2000, 0.001, 0.8, 0.3, 0.5],
            ["QT-156", "下游段", "III", 1.0, 0.04, 1800, 0.001, 0.8, 0.3, 0.5],
        ]
        for row_idx, row_data in enumerate(sample_zones, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws1.cell(row=row_idx, column=col_idx, value=value)
    
    # 调整列宽
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 16
    ws1.column_dimensions['H'].width = 14
    
    # ========== Sheet 2: 逐日流量 ==========
    ws2 = wb.create_sheet("逐日流量")
    
    # 表头
    ws2.cell(row=1, column=1, value="日期").font = Font(bold=True)
    if with_sample:
        zone_ids = ["QT-153", "QT-154", "QT-155", "QT-156"]
        for col, zone_id in enumerate(zone_ids, 2):
            ws2.cell(row=1, column=col, value=zone_id).font = Font(bold=True)
        
        # 生成示例数据
        np.random.seed(42)
        dates = pd.date_range('1992-01-01', '1993-12-31', freq='D')
        for row_idx, date in enumerate(dates, 2):
            ws2.cell(row=row_idx, column=1, value=date.strftime('%Y-%m-%d'))
            for col_idx in range(len(zone_ids)):
                ws2.cell(row=row_idx, column=col_idx + 2, 
                        value=round(np.random.uniform(100, 700), 2))
    
    ws2.column_dimensions['A'].width = 12
    
    # ========== Sheet 3: 水库功能区基础信息 ==========
    ws3 = wb.create_sheet("水库功能区基础信息")
    
    headers3 = ["功能区", "名称", "K(1/s)", "b", "Cs", "C0"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    if with_sample:
        sample_reservoirs = [
            ["SK-01", "青山水库", 0.000002, 0.2, 0.5, 0.02],
            ["SK-02", "碧湖水库", 0.000002, 0.2, 1.0, 0.02],
        ]
        for row_idx, row_data in enumerate(sample_reservoirs, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws3.cell(row=row_idx, column=col_idx, value=value)
    
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 12
    
    # ========== Sheet 4: 水库逐日库容 ==========
    ws4 = wb.create_sheet("水库逐日库容")
    
    ws4.cell(row=1, column=1, value="日期").font = Font(bold=True)
    if with_sample:
        reservoir_ids = ["SK-01", "SK-02"]
        for col, zone_id in enumerate(reservoir_ids, 2):
            ws4.cell(row=1, column=col, value=zone_id).font = Font(bold=True)
        
        # 生成示例数据（水文年：4月→3月）
        dates = pd.date_range('1992-04-01', '1994-03-31', freq='D')
        for row_idx, date in enumerate(dates, 2):
            ws4.cell(row=row_idx, column=1, value=date.strftime('%Y-%m-%d'))
            for col_idx in range(len(reservoir_ids)):
                ws4.cell(row=row_idx, column=col_idx + 2, 
                        value=round(np.random.uniform(40000000, 55000000), 0))
    
    ws4.column_dimensions['A'].width = 12
    
    # ========== 保存 ==========
    # 注意：openpyxl 不支持直接写入 VBA，需要手动导入
    # 这里先保存为 xlsx，然后提示用户手动添加 VBA
    
    wb.save(output_path)
    print(f"✓ 模板已创建: {output_path}")


def print_vba_import_guide():
    """打印 VBA 导入指南"""
    print("""
============================================================
VBA 导入指南（macOS Excel）
============================================================

1. 打开生成的 Excel 文件

2. 按 Option + F11 打开 VBA 编辑器

3. 在左侧「工程资源管理器」中右键点击 VBAProject
   选择「导入文件...」

4. 选择 vba/Module_Main.bas 文件导入

5. 关闭 VBA 编辑器

6. 另存为 .xlsm 格式（启用宏的工作簿）
   - 文件 → 另存为...
   - 格式选择「Excel 宏启用工作簿 (.xlsm)」

7. 添加按钮（可选）：
   - 开发工具 → 按钮（表单控件）
   - 在 Sheet 上画一个按钮
   - 选择「开始计算」宏

提示：如果看不到「开发工具」选项卡，请在 Excel 偏好设置 → 
     功能区和工具栏 → 勾选「开发工具」
============================================================
""")


def main():
    parser = argparse.ArgumentParser(description='创建纳污能力计算 Excel 模板')
    parser.add_argument('--empty', action='store_true', help='创建空模板（不含示例数据）')
    parser.add_argument('--output', '-o', default='纳污能力计算模板.xlsx', 
                        help='输出文件名（默认：纳污能力计算模板.xlsx）')
    
    args = parser.parse_args()
    
    base_dir = Path(__file__).parent
    output_path = base_dir / args.output
    
    print("=" * 60)
    print("创建纳污能力计算 Excel 模板")
    print("=" * 60)
    
    create_template(output_path, with_sample=not args.empty)
    print_vba_import_guide()


if __name__ == '__main__':
    main()
