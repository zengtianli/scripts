#!/usr/bin/env python3
"""
XLSX转CSV转换工具 - 将Excel XLSX文件转换为CSV格式
版本: 2.0.0
作者: tianli
更新: 2024-01-01
"""

import argparse
import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "lib"))
from display import show_error, show_info, show_processing, show_success, show_warning
from file_ops import (
    check_file_extension,
    check_python_packages,
    fatal_error,
    find_files_by_extension,
    show_version_info,
    validate_input_file,
)
from finder import get_input_files
from progress import ProgressTracker

SCRIPT_VERSION = "2.0.0"
SCRIPT_AUTHOR = "tianli"
SCRIPT_UPDATED = "2024-01-01"


def check_dependencies() -> bool:
    show_info("检查依赖项...")
    if not check_python_packages("openpyxl"):
        return False
    show_success("依赖检查完成")
    return True


def convert_xlsx_to_csv_single(
    input_file: Path, output_file: Path | None = None, sheet_name: str | None = None, all_sheets: bool = True
) -> bool:
    try:
        if not validate_input_file(input_file):
            return False

        if not check_file_extension(input_file, "xlsx"):
            show_warning(f"跳过不支持的文件: {input_file.name}")
            return False

        show_processing(f"转换: {input_file.name}")
        from openpyxl import load_workbook

        wb = load_workbook(input_file, read_only=True, data_only=True)

        if sheet_name and sheet_name in wb.sheetnames:
            sheets_to_process = [(sheet_name, wb[sheet_name])]
        elif not all_sheets:
            sheets_to_process = [(wb.active.title, wb.active)]
        else:
            sheets_to_process = [(name, wb[name]) for name in wb.sheetnames]

        success_count = 0
        for name, ws in sheets_to_process:
            if output_file and len(sheets_to_process) == 1:
                current_output = output_file
            else:
                current_output = input_file.parent / f"{input_file.stem}_{name}.csv"

            with open(current_output, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if cell is None else str(cell) for cell in row])

            show_success(f"已转换工作表 '{name}' -> {current_output.name}")
            success_count += 1

        return success_count > 0

    except Exception as e:
        show_error(f"转换失败: {input_file.name} - {e}")
        return False


def batch_process(directory: Path, recursive: bool = False, all_sheets: bool = True) -> None:
    show_info(f"处理目录: {directory}")
    files = find_files_by_extension(directory, "xlsx", recursive)

    if not files:
        show_warning("未找到XLSX文件")
        return

    show_info(f"找到 {len(files)} 个XLSX文件")
    tracker = ProgressTracker()

    for i, file in enumerate(files, 1):
        show_processing(f"进度 ({i}/{len(files)}): {file.name}")
        if convert_xlsx_to_csv_single(file, all_sheets=all_sheets):
            tracker.add_success()
        else:
            tracker.add_failure()

    tracker.show_summary("文件转换")


def show_version() -> None:
    show_version_info(SCRIPT_VERSION, SCRIPT_AUTHOR, SCRIPT_UPDATED)


def show_help() -> None:
    print(f"""
用法: python3 {sys.argv[0]} [选项] [输入] [输出]

参数:
  输入            输入XLSX文件或目录
  输出            输出CSV文件（可选，仅对单文件有效）

选项:
  -r, --recursive   递归处理子目录
  -s, --sheet NAME  指定要转换的工作表名称
  -d, --default     仅转换默认工作表
  -h, --help        显示此帮助信息
  --version         显示版本信息

依赖:
  - openpyxl
    """)


def main():
    parser = argparse.ArgumentParser(description="XLSX转CSV转换工具", add_help=False)
    parser.add_argument("-r", "--recursive", action="store_true", help="递归处理子目录")
    parser.add_argument("-s", "--sheet", help="指定要转换的工作表名称")
    parser.add_argument("-d", "--default", action="store_true", help="仅转换默认工作表")
    parser.add_argument("-h", "--help", action="store_true", help="显示帮助信息")
    parser.add_argument("--version", action="store_true", help="显示版本信息")
    args, unknown = parser.parse_known_args()

    if args.help:
        show_help()
        return

    if args.version:
        show_version()
        return

    if not check_dependencies():
        sys.exit(1)

    all_sheets = not (args.sheet or args.default)

    # 获取输入文件列表
    input_files = [Path(f) for f in get_input_files(unknown, expected_ext="xlsx")]

    if not input_files:
        show_warning("未找到XLSX文件")
        sys.exit(1)

    show_info(f"找到 {len(input_files)} 个XLSX文件")
    tracker = ProgressTracker()

    # 批量处理文件
    for i, file_path in enumerate(input_files, 1):
        show_processing(f"进度 ({i}/{len(input_files)}): {file_path.name}")
        if convert_xlsx_to_csv_single(file_path, sheet_name=args.sheet, all_sheets=all_sheets):
            tracker.add_success()
        else:
            tracker.add_failure()

    tracker.show_summary("文件转换")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        show_warning("用户中断操作")
        sys.exit(1)
    except Exception as e:
        fatal_error(f"程序执行失败: {e}")
