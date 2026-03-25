#!/usr/bin/env python3
"""
Excel工作表分离工具 - 将单个Excel文件按工作表拆分为多个文件
版本: 3.0.0
作者: tianli
更新: 2025-11-11
"""

import argparse
import multiprocessing
import sys
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "lib"))
from display import show_error, show_info, show_processing, show_success, show_warning
from file_ops import (
    check_file_extension,
    check_python_packages,
    fatal_error,
    get_file_basename,
    show_version_info,
    validate_input_file,
)
from finder import get_input_files

SCRIPT_VERSION = "3.0.0"
SCRIPT_AUTHOR = "tianli"
SCRIPT_UPDATED = "2025-11-11"


def check_dependencies() -> bool:
    show_info("检查依赖项...")
    if not check_python_packages("openpyxl"):
        return False
    show_success("依赖检查完成")
    return True


def copy_sheet_with_formatting(source_sheet, target_sheet):
    """
    完整复制工作表，包括所有格式、公式、合并单元格等
    类似 Excel 的 "Move and Copy" 功能
    """
    # 1. 复制所有单元格的数据、公式和样式
    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet[cell.coordinate]

            # 复制值和公式
            if cell.value:
                target_cell.value = cell.value

            # 复制完整的样式
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = copy(cell.number_format)
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)

            # 复制超链接
            if cell.hyperlink:
                target_cell.hyperlink = copy(cell.hyperlink)

            # 复制注释
            if cell.comment:
                target_cell.comment = copy(cell.comment)

    # 2. 复制合并单元格信息
    for merged_cell_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell_range))

    # 3. 复制列宽
    for col_letter, dimension in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_letter].width = dimension.width
        if dimension.hidden:
            target_sheet.column_dimensions[col_letter].hidden = True

    # 4. 复制行高
    for row_num, dimension in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_num].height = dimension.height
        if dimension.hidden:
            target_sheet.row_dimensions[row_num].hidden = True

    # 5. 复制工作表属性
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.page_setup = copy(source_sheet.page_setup)
    target_sheet.print_options = copy(source_sheet.print_options)

    # 6. 复制冻结窗格
    if source_sheet.freeze_panes:
        target_sheet.freeze_panes = source_sheet.freeze_panes

    # 7. 复制筛选器
    if source_sheet.auto_filter:
        target_sheet.auto_filter.ref = source_sheet.auto_filter.ref

    # 8. 复制条件格式
    if hasattr(source_sheet, "conditional_formatting"):
        for range_string, rules in source_sheet.conditional_formatting._cf_rules.items():
            target_sheet.conditional_formatting._cf_rules[range_string] = copy(rules)

    # 9. 复制数据验证
    if hasattr(source_sheet, "data_validations"):
        for dv in source_sheet.data_validations.dataValidation:
            target_sheet.data_validations.append(copy(dv))


def process_single_sheet(args):
    """
    处理单个工作表（用于多进程并行）

    Args:
        args: 元组 (input_file_path, sheet_name, base_name, output_dir)

    Returns:
        元组 (success, sheet_name, output_file_name, error_msg)
    """
    input_file_path, sheet_name, base_name, output_dir = args

    try:
        # 每个进程独立加载源文件
        source_wb = load_workbook(input_file_path, data_only=False, keep_vba=True)

        # 创建新的工作簿
        new_wb = Workbook()
        new_wb.remove(new_wb.active)  # 删除默认的空工作表

        # 创建新工作表
        new_sheet = new_wb.create_sheet(sheet_name)
        source_sheet = source_wb[sheet_name]

        # 完整复制工作表内容和格式
        copy_sheet_with_formatting(source_sheet, new_sheet)

        # 保存新文件
        output_file = Path(output_dir) / f"{base_name}_{sheet_name}.xlsx"
        new_wb.save(output_file)

        return (True, sheet_name, output_file.name, None)

    except Exception as e:
        return (False, sheet_name, None, str(e))


def split_excel_file(input_file: Path, max_workers: int = None) -> bool:
    """
    并行拆分Excel文件

    Args:
        input_file: 输入的Excel文件路径
        max_workers: 最大并行工作进程数，None表示使用CPU核心数

    Returns:
        是否成功
    """
    try:
        if not validate_input_file(input_file):
            return False

        if not check_file_extension(input_file, "xlsx"):
            show_warning(f"跳过非XLSX文件: {input_file.name}")
            return False

        show_processing(f"正在读取Excel文件: {input_file.name}")

        # 快速读取工作表名称（不加载全部数据）
        source_wb = load_workbook(input_file, data_only=False, read_only=True)
        sheet_names = source_wb.sheetnames
        source_wb.close()

        if not sheet_names:
            show_warning(f"文件 '{input_file.name}' 中没有找到工作表。")
            return True

        total_sheets = len(sheet_names)
        show_info(f"找到 {total_sheets} 个工作表: {', '.join(sheet_names)}")

        # 确定并行工作进程数
        if max_workers is None:
            max_workers = min(multiprocessing.cpu_count(), total_sheets)

        show_info(f"使用 {max_workers} 个并行进程加速处理...")

        base_name = get_file_basename(input_file)
        output_dir = str(input_file.parent)

        # 准备所有任务参数
        tasks = [(str(input_file), sheet_name, base_name, output_dir) for sheet_name in sheet_names]

        # 并行处理所有工作表
        start_time = time.time()
        success_count = 0
        failed_sheets = []

        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            # 提交所有任务
            future_to_sheet = {executor.submit(process_single_sheet, task): task[1] for task in tasks}

            # 收集结果（按完成顺序）
            for future in as_completed(future_to_sheet):
                success, sheet_name, output_file_name, error_msg = future.result()

                if success:
                    success_count += 1
                    show_success(
                        f"[{success_count}/{total_sheets}] 已保存工作表 '{sheet_name}' 到 '{output_file_name}'"
                    )
                else:
                    failed_sheets.append((sheet_name, error_msg))
                    show_error(
                        f"[{success_count + len(failed_sheets)}/{total_sheets}] "
                        f"处理工作表 '{sheet_name}' 失败: {error_msg}"
                    )

        elapsed_time = time.time() - start_time

        # 显示总结
        show_info(f"处理完成，用时: {elapsed_time:.2f} 秒")
        show_success(f"成功: {success_count}/{total_sheets} 个工作表")

        if failed_sheets:
            show_warning(f"失败: {len(failed_sheets)}/{total_sheets} 个工作表")
            for sheet_name, error_msg in failed_sheets:
                show_error(f"  - {sheet_name}: {error_msg}")
            return False

        return True

    except Exception as e:
        show_error(f"处理文件 '{input_file.name}' 时发生错误: {e}")
        import traceback

        show_error(f"详细错误: {traceback.format_exc()}")
        return False


def show_version() -> None:
    show_version_info(SCRIPT_VERSION, SCRIPT_AUTHOR, SCRIPT_UPDATED)


def show_help() -> None:
    print(f"""
Excel工作表分离工具 - 将单个Excel文件按工作表拆分为多个文件

用法:
    python3 {sys.argv[0]} [选项] <输入文件>

参数:
    输入文件         要拆分的Excel文件 (.xlsx)

选项:
    -h, --help       显示此帮助信息
    --version        显示版本信息
    -j N, --jobs N   并行处理的进程数（默认: CPU核心数）

示例:
    python3 {sys.argv[0]} data.xlsx
    python3 {sys.argv[0]} data.xlsx -j 4        # 使用4个并行进程
    python3 {sys.argv[0]} data.xlsx --jobs 8    # 使用8个并行进程

功能:
    - 将一个包含多个工作表的Excel文件拆分为多个单独的Excel文件
    - 每个新文件以原文件名和工作表名命名
    - 完整保留所有格式、公式、样式（类似 Excel 的 "Move and Copy" 功能）
    - ⚡ 多进程并行处理，大幅提升处理速度
    
保留内容:
    ✓ 单元格数据和公式
    ✓ 字体、颜色、边框、对齐等样式
    ✓ 合并单元格
    ✓ 列宽和行高
    ✓ 冻结窗格
    ✓ 筛选器
    ✓ 条件格式
    ✓ 数据验证
    ✓ 超链接和注释

性能:
    - 自动使用多核CPU并行处理
    - 处理速度提升约 N 倍（N = CPU核心数）
    - 显示实时进度和处理时间

依赖:
    - openpyxl
    """)


def main():
    # 无参数时从 Finder 获取选中的文件
    if len(sys.argv) == 1:
        files = get_input_files([], expected_ext="xlsx")
        if files:
            sys.argv.extend(files)

    parser = argparse.ArgumentParser(description="Excel工作表分离工具", add_help=False)

    parser.add_argument("input_file", nargs="?", help="要拆分的Excel文件")
    parser.add_argument("-h", "--help", action="store_true", help="显示帮助信息")
    parser.add_argument("--version", action="store_true", help="显示版本信息")
    parser.add_argument("-j", "--jobs", type=int, default=None, metavar="N", help="并行处理的进程数（默认: CPU核心数）")

    args = parser.parse_args()

    if args.help:
        show_help()
        return

    if args.version:
        show_version()
        return

    if not args.input_file:
        show_help()
        fatal_error("错误: 未提供输入文件。")

    if not check_dependencies():
        sys.exit(1)

    # 验证并行进程数
    max_workers = args.jobs
    if max_workers is not None and max_workers < 1:
        fatal_error("错误: 并行进程数必须大于等于1。")

    input_path = Path(args.input_file)

    if not split_excel_file(input_path, max_workers=max_workers):
        sys.exit(1)

    show_success("所有操作完成。")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        show_warning("用户中断操作")
        sys.exit(1)
    except Exception as e:
        fatal_error(f"程序执行失败: {e}")
