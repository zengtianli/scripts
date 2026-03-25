#!/usr/bin/env python3
"""
数据格式转换统一工具 - 合并 8 个格式互转脚本
用法: python3 convert.py <子命令> [选项] [输入文件...]

子命令:
  csv-from-txt    TXT → CSV（空白符分隔转逗号分隔）
  csv-to-txt      CSV → TXT（逗号分隔转制表符分隔）
  csv-merge-txt   合并多个 TXT 文件为一个 CSV（按列）
  xlsx-from-csv   CSV → XLSX
  xlsx-from-txt   TXT → XLSX
  xlsx-from-xls   XLS → XLSX（旧格式转新格式）
  xlsx-to-csv     XLSX → CSV（支持多工作表）
  xlsx-to-txt     XLSX → TXT（支持多工作表）

版本: 3.0.0
作者: tianli
"""

import argparse
import csv
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "lib"))
from display import show_error, show_info, show_processing, show_success, show_warning
from file_ops import (
    check_file_extension,
    check_python_packages,
    fatal_error,
    find_files_by_extension,
    show_version_info,
    validate_input_file,
)
from finder import get_finder_current_dir, get_input_files
from progress import ProgressTracker

SCRIPT_VERSION = "3.0.0"
SCRIPT_AUTHOR = "tianli"
SCRIPT_UPDATED = "2026-03-25"


# ── 转换函数 ──────────────────────────────────────────────────


def _csv_from_txt(input_file: Path, output_file: Path | None = None, **_kw) -> bool:
    """TXT → CSV：空白符分隔转逗号分隔"""
    if not validate_input_file(input_file) or not check_file_extension(input_file, "txt"):
        return False
    if output_file is None:
        output_file = input_file.with_suffix(".csv")
    show_processing(f"转换: {input_file.name} -> {output_file.name}")
    with open(input_file, encoding="utf-8") as f:
        content = f.readlines()
    rows = []
    for line in content:
        line = line.strip()
        if line:
            rows.append(re.sub(r"\s+", ",", line).split(","))
    with open(output_file, "w", encoding="utf-8", newline="") as f:
        csv.writer(f).writerows(rows)
    show_success(f"转换完成: {output_file.name}")
    return True


def _csv_to_txt(input_file: Path, output_file: Path | None = None, **_kw) -> bool:
    """CSV → TXT：逗号分隔转制表符分隔"""
    if not validate_input_file(input_file) or not check_file_extension(input_file, "csv"):
        return False
    if output_file is None:
        output_file = input_file.with_suffix(".txt")
    show_processing(f"转换: {input_file.name} -> {output_file.name}")
    with open(input_file, encoding="utf-8") as f_in, open(output_file, "w", encoding="utf-8") as f_out:
        for row in csv.reader(f_in):
            f_out.write("\t".join(row) + "\n")
    show_success(f"转换完成: {output_file.name}")
    return True


def _xlsx_from_csv(input_file: Path, output_file: Path | None = None, **_kw) -> bool:
    """CSV → XLSX"""
    if not validate_input_file(input_file) or not check_file_extension(input_file, "csv"):
        return False
    if output_file is None:
        output_file = input_file.with_suffix(".xlsx")
    show_processing(f"转换: {input_file.name} -> {output_file.name}")
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    with open(input_file, encoding="utf-8") as f:
        for row in csv.reader(f):
            ws.append(row)
    wb.save(output_file)
    show_success(f"转换完成: {output_file.name}")
    return True


def _xlsx_from_txt(input_file: Path, output_file: Path | None = None, **_kw) -> bool:
    """TXT → XLSX（制表符分隔）"""
    if not validate_input_file(input_file) or not check_file_extension(input_file, "txt"):
        return False
    if output_file is None:
        output_file = input_file.with_suffix(".xlsx")
    show_processing(f"转换: {input_file.name} -> {output_file.name}")
    import pandas as pd

    df = pd.read_csv(input_file, sep="\t", encoding="utf-8", engine="python")
    df.to_excel(output_file, index=False)
    show_success(f"转换完成: {output_file.name}")
    return True


def _xlsx_from_xls(input_file: Path, output_file: Path | None = None, **_kw) -> bool:
    """XLS → XLSX（旧格式转新格式）"""
    if output_file is None:
        output_file = input_file.with_suffix(".xlsx")
    show_processing(f"转换: {input_file.name} -> {output_file.name}")
    import xlrd
    from openpyxl import Workbook

    xls_book = xlrd.open_workbook(str(input_file))
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name in xls_book.sheet_names():
        xls_sheet = xls_book.sheet_by_name(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=xls_sheet.cell_value(row_idx, col_idx))
    wb.save(str(output_file))
    show_success(f"转换完成: {output_file.name}")
    return True


def _xlsx_to_csv(
    input_file: Path, output_file: Path | None = None, sheet_name: str | None = None, all_sheets: bool = True, **_kw
) -> bool:
    """XLSX → CSV（支持多工作表）"""
    if not validate_input_file(input_file) or not check_file_extension(input_file, "xlsx"):
        return False
    show_processing(f"转换: {input_file.name}")
    from openpyxl import load_workbook

    wb = load_workbook(input_file, read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        sheets = [(sheet_name, wb[sheet_name])]
    elif not all_sheets:
        sheets = [(wb.active.title, wb.active)]
    else:
        sheets = [(name, wb[name]) for name in wb.sheetnames]
    for name, ws in sheets:
        if output_file and len(sheets) == 1:
            out = output_file
        else:
            out = input_file.parent / f"{input_file.stem}_{name}.csv"
        with open(out, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if c is None else str(c) for c in row])
        show_success(f"工作表 '{name}' -> {out.name}")
    return True


def _xlsx_to_txt(input_file: Path, output_file: Path | None = None, **_kw) -> bool:
    """XLSX → TXT（制表符分隔，支持多工作表）"""
    if not validate_input_file(input_file) or not check_file_extension(input_file, "xlsx"):
        return False
    show_processing(f"转换: {input_file.name}")
    import pandas as pd

    excel_file = pd.ExcelFile(input_file)
    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(input_file, sheet_name=sheet_name)
        if output_file and len(excel_file.sheet_names) == 1:
            out = output_file
        else:
            out = input_file.parent / f"{input_file.stem}_{sheet_name}.txt"
        df.to_csv(out, sep="\t", index=False)
        show_success(f"工作表 '{sheet_name}' -> {out.name}")
    return True


def _csv_merge_txt(target_dir: Path, output_file: Path | None = None, **_kw) -> bool:
    """合并目录中所有 TXT 文件为一个 CSV（按列拼接）"""
    if output_file is None:
        output_file = target_dir / "merged.csv"
    txt_files = sorted(target_dir.glob("*.txt"))
    if not txt_files:
        show_error(f"目录 '{target_dir}' 中未找到 .txt 文件")
        return False

    def extract_number(fp: Path):
        match = re.match(r"(\d+)(_\d+)?\.txt", fp.name)
        return (int(match.group(1)), match.group(2) or "") if match else (float("inf"), "")

    txt_files.sort(key=extract_number)
    show_info(f"找到 {len(txt_files)} 个 TXT 文件")
    max_lines = max(sum(1 for _ in open(f, encoding="utf-8")) for f in txt_files)
    if max_lines == 0:
        output_file.touch()
        show_success("所有文件为空，生成空 CSV")
        return True
    lines = [[] for _ in range(max_lines)]
    for f in txt_files:
        show_info(f"处理: {f.name}")
        file_lines = [line.strip() for line in open(f, encoding="utf-8")]
        for i in range(max_lines):
            lines[i].append(file_lines[i] if i < len(file_lines) else "")
    with open(output_file, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(lines)
    show_success(f"合并完成: {output_file.name}")
    return True


# ── 子命令注册表 ─────────────────────────────────────────────

CONVERTERS = {
    "csv-from-txt": {"fn": _csv_from_txt, "src_ext": "txt", "deps": []},
    "csv-to-txt": {"fn": _csv_to_txt, "src_ext": "csv", "deps": []},
    "csv-merge-txt": {"fn": _csv_merge_txt, "special": True, "deps": []},
    "xlsx-from-csv": {"fn": _xlsx_from_csv, "src_ext": "csv", "deps": ["openpyxl"]},
    "xlsx-from-txt": {"fn": _xlsx_from_txt, "src_ext": "txt", "deps": ["pandas", "openpyxl"]},
    "xlsx-from-xls": {"fn": _xlsx_from_xls, "src_ext": "xls", "deps": ["xlrd", "openpyxl"]},
    "xlsx-to-csv": {"fn": _xlsx_to_csv, "src_ext": "xlsx", "deps": ["openpyxl"]},
    "xlsx-to-txt": {"fn": _xlsx_to_txt, "src_ext": "xlsx", "deps": ["pandas", "openpyxl"]},
}


# ── 主入口 ───────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="数据格式转换统一工具",
        usage="python3 convert.py <子命令> [选项] [输入文件...]",
        add_help=False,
    )
    parser.add_argument("command", nargs="?", help="转换子命令")
    parser.add_argument("-r", "--recursive", action="store_true", help="递归处理子目录")
    parser.add_argument("-s", "--sheet", help="指定工作表名称 (xlsx-to-csv)")
    parser.add_argument("-d", "--default-sheet", action="store_true", help="仅转换默认工作表 (xlsx-to-csv)")
    parser.add_argument("-h", "--help", action="store_true", help="显示帮助信息")
    parser.add_argument("--version", action="store_true", help="显示版本信息")
    args, unknown = parser.parse_known_args()

    if args.version:
        show_version_info(SCRIPT_VERSION, SCRIPT_AUTHOR, SCRIPT_UPDATED)
        return

    if args.help or not args.command:
        print(__doc__)
        return

    cmd = args.command
    if cmd not in CONVERTERS:
        show_error(f"未知子命令: {cmd}")
        print(f"可用子命令: {', '.join(CONVERTERS.keys())}")
        sys.exit(1)

    conv = CONVERTERS[cmd]

    # 检查依赖
    if conv.get("deps") and not check_python_packages(*conv["deps"]):
        sys.exit(1)

    # csv-merge-txt 特殊处理：输入是目录
    if conv.get("special"):
        if unknown:
            target_dir = Path(unknown[0])
        else:
            finder_dir = get_finder_current_dir()
            target_dir = Path(finder_dir) if finder_dir else Path(".")
        output_file = Path(unknown[1]) if len(unknown) > 1 else None
        if not target_dir.is_dir():
            fatal_error(f"不是有效目录: {target_dir}")
        conv["fn"](target_dir, output_file)
        return

    # 标准转换流程
    src_ext = conv["src_ext"]
    input_files = [Path(f) for f in get_input_files(unknown, expected_ext=src_ext)]

    if not input_files:
        show_warning(f"未找到 {src_ext.upper()} 文件")
        sys.exit(1)

    show_info(f"找到 {len(input_files)} 个 {src_ext.upper()} 文件")
    tracker = ProgressTracker()

    extra_kw = {}
    if cmd == "xlsx-to-csv":
        extra_kw["sheet_name"] = args.sheet
        extra_kw["all_sheets"] = not (args.sheet or args.default_sheet)

    for i, fp in enumerate(input_files, 1):
        show_processing(f"进度 ({i}/{len(input_files)}): {fp.name}")
        try:
            if conv["fn"](fp, **extra_kw):
                tracker.add_success()
            else:
                tracker.add_failure()
        except Exception as e:
            show_error(f"转换失败: {fp.name} - {e}")
            tracker.add_failure()

    tracker.show_summary("文件转换")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        show_warning("用户中断操作")
        sys.exit(1)
    except Exception as e:
        fatal_error(f"程序执行失败: {e}")
