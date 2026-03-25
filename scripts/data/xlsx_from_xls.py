#!/usr/bin/env python3
"""
将 .xls 文件转换为 .xlsx 格式
"""

import sys
from pathlib import Path

import xlrd
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "lib"))
from display import show_error, show_info, show_success
from finder import get_input_files


def convert_xls_to_xlsx(xls_path: str) -> bool:
    """将单个 .xls 文件转换为 .xlsx"""
    input_path = Path(xls_path)
    output_path = input_path.with_suffix(".xlsx")

    try:
        xls_book = xlrd.open_workbook(xls_path)
        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name in xls_book.sheet_names():
            xls_sheet = xls_book.sheet_by_name(sheet_name)
            ws = wb.create_sheet(title=sheet_name)

            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    ws.cell(row=row_idx + 1, column=col_idx + 1, value=xls_sheet.cell_value(row_idx, col_idx))

        wb.save(str(output_path))
        show_success(f"{input_path.name} -> {output_path.name}")
        return True
    except Exception as e:
        show_error(f"{input_path.name}: {e}")
        return False


if __name__ == "__main__":
    files = get_input_files(sys.argv[1:], expected_ext="xls")
    if not files:
        sys.exit(1)

    success = sum(1 for f in files if convert_xls_to_xlsx(f))
    show_info(f"完成：{success}/{len(files)} 个文件转换成功")
