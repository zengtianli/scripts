#!/usr/bin/env python3
"""
Excel 核心功能模块
提供所有 xlsx 相关的转换和处理功能
"""

from pathlib import Path

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


def xlsx_to_csv(input_file: Path, output_file: Path | None = None,
                sheet: str | None = None) -> bool:
    """Excel 转 CSV"""
    if not HAS_PANDAS:
        return False

    if output_file is None:
        output_file = input_file.with_suffix('.csv')

    try:
        df = pd.read_excel(input_file, sheet_name=sheet or 0)
        df.to_csv(output_file, index=False, encoding='utf-8')
        return True
    except Exception:
        return False


def xlsx_to_txt(input_file: Path, output_file: Path | None = None,
                delimiter: str = '\t') -> bool:
    """Excel 转 TXT"""
    if not HAS_PANDAS:
        return False

    if output_file is None:
        output_file = input_file.with_suffix('.txt')

    try:
        df = pd.read_excel(input_file)
        df.to_csv(output_file, sep=delimiter, index=False, encoding='utf-8')
        return True
    except Exception:
        return False


def csv_to_xlsx(input_file: Path, output_file: Path | None = None) -> bool:
    """CSV 转 Excel"""
    if not HAS_PANDAS:
        return False

    if output_file is None:
        output_file = input_file.with_suffix('.xlsx')

    try:
        df = pd.read_csv(input_file, encoding='utf-8')
        df.to_excel(output_file, index=False)
        return True
    except Exception:
        return False


def txt_to_xlsx(input_file: Path, output_file: Path | None = None,
                delimiter: str = '\t') -> bool:
    """TXT 转 Excel"""
    if not HAS_PANDAS:
        return False

    if output_file is None:
        output_file = input_file.with_suffix('.xlsx')

    try:
        df = pd.read_csv(input_file, sep=delimiter, encoding='utf-8', engine='python')
        df.to_excel(output_file, index=False)
        return True
    except Exception:
        return False


def xls_to_xlsx(input_file: Path, output_file: Path | None = None) -> bool:
    """旧版 .xls 转 .xlsx"""
    if not HAS_PANDAS:
        return False

    if output_file is None:
        output_file = input_file.with_suffix('.xlsx')

    try:
        df = pd.read_excel(input_file, engine='xlrd')
        df.to_excel(output_file, index=False)
        return True
    except Exception:
        return False


def split_sheets(input_file: Path, output_dir: Path | None = None) -> list[Path]:
    """将 Excel 的每个 sheet 拆分为单独文件"""
    if not HAS_PANDAS:
        return []

    if output_dir is None:
        output_dir = input_file.parent / f"{input_file.stem}_sheets"
    output_dir.mkdir(exist_ok=True)

    try:
        excel = pd.ExcelFile(input_file)
        created_files = []

        for sheet_name in excel.sheet_names:
            df = pd.read_excel(excel, sheet_name=sheet_name)
            safe_name = "".join(c if c.isalnum() or c in '._- ' else '_' for c in sheet_name)
            output_file = output_dir / f"{safe_name}.xlsx"
            df.to_excel(output_file, index=False)
            created_files.append(output_file)

        return created_files
    except Exception:
        return []


def merge_tables(input_files: list[Path], output_file: Path,
                 key_column: str | None = None) -> bool:
    """合并多个 Excel 表格"""
    if not HAS_PANDAS:
        return False

    try:
        dfs = [pd.read_excel(f) for f in input_files]

        if key_column and all(key_column in df.columns for df in dfs):
            # 按键列合并
            result = dfs[0]
            for df in dfs[1:]:
                result = pd.merge(result, df, on=key_column, how='outer')
        else:
            # 简单拼接
            result = pd.concat(dfs, ignore_index=True)

        result.to_excel(output_file, index=False)
        return True
    except Exception:
        return False


def lowercase_headers(input_file: Path, output_file: Path | None = None) -> bool:
    """将列名转为小写"""
    if not HAS_PANDAS:
        return False

    if output_file is None:
        output_file = input_file.parent / f"{input_file.stem}_lowercase.xlsx"

    try:
        df = pd.read_excel(input_file)
        df.columns = [str(c).lower() for c in df.columns]
        df.to_excel(output_file, index=False)
        return True
    except Exception:
        return False


def read_xlsx(input_file: Path, sheet: str | None = None) -> str:
    """读取 Excel 内容为文本"""
    if not HAS_PANDAS:
        return "需要安装 pandas"

    try:
        excel = pd.ExcelFile(input_file)
        content = [f"文件: {input_file.name}", f"共 {len(excel.sheet_names)} 个 sheet", "=" * 60]

        sheets_to_read = [sheet] if sheet else excel.sheet_names

        for sheet_name in sheets_to_read:
            if sheet_name not in excel.sheet_names:
                continue
            df = pd.read_excel(excel, sheet_name=sheet_name)
            content.append(f"\n=== Sheet: {sheet_name} ===")
            content.append(f"维度: {df.shape[0]} 行 × {df.shape[1]} 列")
            content.append(df.to_string(index=True, max_rows=100))

        return "\n".join(content)
    except Exception as e:
        return f"读取失败: {e}"


def xls_to_xlsx(input_path: Path) -> bool:
    """将 xls 转换为 xlsx"""
    try:
        import xlrd
        from openpyxl import Workbook
        wb_old = xlrd.open_workbook(str(input_path))
        wb_new = Workbook()
        for sheet_name in wb_old.sheet_names():
            ws_old = wb_old.sheet_by_name(sheet_name)
            ws_new = wb_new.create_sheet(title=sheet_name)
            for row in range(ws_old.nrows):
                for col in range(ws_old.ncols):
                    ws_new.cell(row=row+1, column=col+1, value=ws_old.cell_value(row, col))
        if 'Sheet' in wb_new.sheetnames:
            del wb_new['Sheet']
        wb_new.save(str(input_path.with_suffix('.xlsx')))
        return True
    except Exception as e:
        print(f"错误: {e}")
        return False

def split_sheets(input_path: Path) -> list:
    """将工作簿拆分为多个文件"""
    try:
        from openpyxl import Workbook, load_workbook
        wb = load_workbook(str(input_path))
        created = []
        out_dir = input_path.parent / f"{input_path.stem}_split"
        out_dir.mkdir(exist_ok=True)
        for sheet_name in wb.sheetnames:
            wb_new = Workbook()
            ws_new = wb_new.active
            ws_new.title = sheet_name
            ws_old = wb[sheet_name]
            for row in ws_old.iter_rows():
                for cell in row:
                    ws_new.cell(row=cell.row, column=cell.column, value=cell.value)
            out_path = out_dir / f"{sheet_name}.xlsx"
            wb_new.save(str(out_path))
            created.append(out_path)
        return created
    except Exception as e:
        print(f"错误: {e}")
        return []

def merge_tables(files: list, output_path: Path) -> bool:
    """合并多个 xlsx 文件"""
    try:
        from openpyxl import Workbook, load_workbook
        wb_out = Workbook()
        ws_out = wb_out.active
        row_offset = 0
        for i, f in enumerate(files):
            wb = load_workbook(str(f))
            ws = wb.active
            for row in ws.iter_rows():
                for cell in row:
                    ws_out.cell(row=cell.row + row_offset, column=cell.column, value=cell.value)
            row_offset = ws_out.max_row
        wb_out.save(str(output_path))
        return True
    except Exception as e:
        print(f"错误: {e}")
        return False

def read_xlsx(input_path: Path) -> str:
    """读取 xlsx 内容为文本"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(input_path))
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"\n=== {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                lines.append('\t'.join(str(c) if c else '' for c in row))
        return '\n'.join(lines)
    except Exception as e:
        return f"错误: {e}"

def lowercase_headers(input_path: Path) -> bool:
    """将表头转为小写"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(input_path))
        for ws in wb.worksheets:
            for cell in ws[1]:
                if cell.value and isinstance(cell.value, str):
                    cell.value = cell.value.lower()
        wb.save(str(input_path))
        return True
    except Exception as e:
        print(f"错误: {e}")
        return False
